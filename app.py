import io
import csv
import base64
import openpyxl
from flask import Flask, request, jsonify

# Initialize the Flask application
app = Flask(__name__)

# --- Helper Functions ---

def load_workbook_from_b64(b64_string):
    """Decodes a base64 string and loads it into an openpyxl workbook."""
    # The MCP client sends a data URL: "data:application/vnd.ms-excel;base64,UEsDBBQ..."
    # We need to strip the prefix to get the pure base64 data.
    try:
        header, encoded = b64_string.split(",", 1)
        file_bytes = base64.b64decode(encoded)
        return openpyxl.load_workbook(filename=io.BytesIO(file_bytes))
    except (ValueError, TypeError, base64.binascii.Error) as e:
        # If splitting or decoding fails, it's a bad request.
        raise ValueError(f"Invalid base64 file format: {e}")
    except Exception as e:
        # If openpyxl can't load the file, it's likely not a valid Excel file.
        raise IOError(f"Failed to parse Excel file: {e}")

def parse_range_string(range_str):
    """Parses a string like 'Sheet1!A1:B10' into sheet name and cell range."""
    if '!' not in range_str:
        raise ValueError("Invalid range format. Expected 'SheetName!A1:B10'.")
    return range_str.split('!', 1)

# --- MCP Endpoints ---

# --- Health Check Endpoint ---
@app.route('/healthz', methods=['GET'])
def healthz():
    """Simple health check endpoint."""
    return jsonify({"status": "ok"}), 200

@app.route('/mcp/info', methods=['GET'])
def get_info():
    """
    This endpoint returns the server's metadata.
    It describes the server and lists all the tools it provides.
    """
    server_info = {
        "name": "MS Excel Tools",
        "description": "A set of tools to perform basic read and write operations on Microsoft Excel (.xlsx) files.",
        "tools": [
            {
                "id": "sum_range",
                "name": "Sum Cell Range",
                "description": "Calculates the sum of all numbers in a given cell range (e.g., 'Sheet1!A1:A10').",
                "parameters": [
                    {"name": "file", "type": "file", "description": "The .xlsx file to process."},
                    {"name": "range", "type": "string", "description": "The cell range string (e.g., 'Sheet1!A1:A10')."}
                ]
            },
            {
                "id": "avg_range",
                "name": "Average Cell Range",
                "description": "Calculates the average of all numbers in a given cell range.",
                "parameters": [
                    {"name": "file", "type": "file", "description": "The .xlsx file to process."},
                    {"name": "range", "type": "string", "description": "The cell range string (e.g., 'Sheet1!A1:A10')."}
                ]
            },
            {
                "id": "get_cell",
                "name": "Get Cell Value",
                "description": "Retrieves the value from a single, specific cell (e.g., 'Sheet1!B2').",
                "parameters": [
                    {"name": "file", "type": "file", "description": "The .xlsx file to process."},
                    {"name": "cell", "type": "string", "description": "The cell address (e.g., 'Sheet1!B2')."}
                ]
            },
            {
                "id": "set_cell",
                "name": "Set Cell Value",
                "description": "Writes a new value to a specific cell and returns the modified Excel file.",
                "parameters": [
                    {"name": "file", "type": "file", "description": "The .xlsx file to process."},
                    {"name": "cell", "type": "string", "description": "The cell address to modify (e.g., 'Sheet1!C3')."},
                    {"name": "value", "type": "string", "description": "The new value to write into the cell."}
                ]
            },
            {
                "id": "to_csv",
                "name": "Convert to CSV",
                "description": "Converts the first worksheet of an Excel file into CSV format.",
                "parameters": [
                    {"name": "file", "type": "file", "description": "The .xlsx file to convert."}
                ]
            }
        ]
    }
    return jsonify(server_info)


@app.route('/mcp/run', methods=['POST'])
def run_tool():
    """
    This is the main execution endpoint. It receives a tool ID and its parameters,
    runs the corresponding logic, and returns the result.
    """
    try:
        data = request.json
        tool_id = data.get('tool_id')
        params = data.get('parameters', {})
        
        # --- Tool: Sum Range ---
        if tool_id == 'sum_range':
            workbook = load_workbook_from_b64(params['file'])
            sheet_name, cell_range = parse_range_string(params['range'])
            sheet = workbook[sheet_name]
            total = 0
            for row in sheet[cell_range]:
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        total += cell.value
            return jsonify({"result": {"value": total}})

        # --- Tool: Average Range ---
        elif tool_id == 'avg_range':
            workbook = load_workbook_from_b64(params['file'])
            sheet_name, cell_range = parse_range_string(params['range'])
            sheet = workbook[sheet_name]
            total = 0
            count = 0
            for row in sheet[cell_range]:
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        total += cell.value
                        count += 1
            average = total / count if count > 0 else 0
            return jsonify({"result": {"value": average}})

        # --- Tool: Get Cell Value ---
        elif tool_id == 'get_cell':
            workbook = load_workbook_from_b64(params['file'])
            sheet_name, cell_address = parse_range_string(params['cell'])
            sheet = workbook[sheet_name]
            value = sheet[cell_address].value
            return jsonify({"result": {"value": value}})

        # --- Tool: Set Cell Value (returns a modified file) ---
        elif tool_id == 'set_cell':
            workbook = load_workbook_from_b64(params['file'])
            sheet_name, cell_address = parse_range_string(params['cell'])
            new_value = params['value']
            sheet = workbook[sheet_name]
            
            # Try to convert value to a number if possible
            try:
                new_value = float(new_value)
            except ValueError:
                pass # Keep as string if conversion fails
                
            sheet[cell_address] = new_value

            # Save the modified workbook to an in-memory stream
            file_stream = io.BytesIO()
            workbook.save(file_stream)
            file_stream.seek(0)
            
            # Encode the new file content to base64
            encoded_file = base64.b64encode(file_stream.read()).decode('utf-8')
            # The data URL prefix is important for the client to recognize the file type
            data_url = f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{encoded_file}"
            
            return jsonify({"result": {"file": data_url}})

        # --- Tool: Convert to CSV (returns a new file) ---
        elif tool_id == 'to_csv':
            workbook = load_workbook_from_b64(params['file'])
            sheet = workbook.active  # Use the first/active sheet
            
            # Use io.StringIO to capture CSV output in memory
            string_stream = io.StringIO()
            csv_writer = csv.writer(string_stream)
            
            for row in sheet.iter_rows(values_only=True):
                csv_writer.writerow(row)
            
            # Get the CSV content and encode it
            csv_content = string_stream.getvalue()
            encoded_csv = base64.b64encode(csv_content.encode('utf-8')).decode('utf-8')
            data_url = f"data:text/csv;base64,{encoded_csv}"

            return jsonify({"result": {"file": data_url}})

        # --- If Tool ID is not found ---
        else:
            return jsonify({"error": f"Tool with id '{tool_id}' not found."}), 404

    except (KeyError, TypeError) as e:
        return jsonify({"error": f"Missing or invalid parameter: {e}"}), 400
    except Exception as e:
        # Catch-all for any other errors during execution
        return jsonify({"error": str(e)}), 500


# --- Main execution ---
if __name__ == '__main__':
    # MCP servers must run on port 7777 inside the container
    # Host '0.0.0.0' makes it accessible from outside the container
    app.run(host='0.0.0.0', port=7777)